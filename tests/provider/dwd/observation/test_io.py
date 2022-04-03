# -*- coding: utf-8 -*-
# Copyright (c) 2018-2021, earthobservations developers.
# Distributed under the MIT License. See LICENSE for more info.
import datetime
import json
import os
import shutil
import sqlite3

import dateutil.parser
import mock
import numpy as np
import openpyxl
import pandas as pd
import pyarrow.feather as feather
import pyarrow.parquet as pq
import pytest
import zarr
from surrogate import surrogate

from wetterdienst.core.process import filter_by_date_and_resolution
from wetterdienst.core.scalar.export import ExportMixin
from wetterdienst.core.scalar.result import StationsResult
from wetterdienst.metadata.resolution import Resolution
from wetterdienst.provider.dwd.observation import (
    DwdObservationDataset,
    DwdObservationPeriod,
    DwdObservationRequest,
    DwdObservationResolution,
)
from wetterdienst.settings import Settings

df_station = pd.DataFrame.from_dict(
    {
        "station_id": ["19087"],
        "from_date": [dateutil.parser.isoparse("1957-05-01T00:00:00.000Z")],
        "to_date": [dateutil.parser.isoparse("1995-11-30T00:00:00.000Z")],
        "height": [645.0],
        "latitude": [48.8049],
        "longitude": [13.5528],
        "name": ["Freyung vorm Wald"],
        "state": ["Bayern"],
        "has_file": [False],
    }
)

df_data = pd.DataFrame.from_dict(
    {
        "station_id": ["01048"],
        "dataset": ["climate_summary"],
        "parameter": ["temperature_air_max_200"],
        "date": [dateutil.parser.isoparse("2019-12-28T00:00:00.000Z")],
        "value": [1.3],
        "quality": [None],
    }
)


def test_to_dict():
    """Test export of DataFrame to dictioanry"""
    data = ExportMixin(df=df_data).to_dict()

    assert data == [
        {
            "dataset": "climate_summary",
            "date": "2019-12-28T00:00:00+00:00",
            "parameter": "temperature_air_max_200",
            "quality": None,
            "station_id": "01048",
            "value": 1.3,
        },
    ]


def test_filter_by_date():
    """Test filter by date"""
    df = filter_by_date_and_resolution(df_data, "2019-12-28", Resolution.HOURLY)
    assert not df.empty

    df = filter_by_date_and_resolution(df_data, "2019-12-27", Resolution.HOURLY)
    assert df.empty


def test_filter_by_date_interval():
    """Test filter by date interval"""
    df = filter_by_date_and_resolution(df_data, "2019-12-27/2019-12-29", Resolution.HOURLY)
    assert not df.empty

    df = filter_by_date_and_resolution(df_data, "2020/2022", Resolution.HOURLY)
    assert df.empty


def test_filter_by_date_monthly():
    """Test filter by date in monthly scope"""
    result = pd.DataFrame.from_dict(
        {
            "station_id": ["01048"],
            "dataset": ["climate_summary"],
            "parameter": ["temperature_air_max_200"],
            "from_date": [dateutil.parser.isoparse("2019-12-28T00:00:00.000Z")],
            "to_date": [dateutil.parser.isoparse("2020-01-28T00:00:00.000Z")],
            "value": [1.3],
            "quality": [None],
        }
    )

    df = filter_by_date_and_resolution(result, "2019-12/2020-01", Resolution.MONTHLY)
    assert not df.empty

    df = filter_by_date_and_resolution(result, "2020/2022", Resolution.MONTHLY)
    assert df.empty

    df = filter_by_date_and_resolution(result, "2020", Resolution.MONTHLY)
    assert df.empty


def test_filter_by_date_annual():
    """Test filter by date in annual scope"""
    df = pd.DataFrame.from_dict(
        {
            "station_id": ["01048"],
            "dataset": ["climate_summary"],
            "parameter": ["temperature_air_max_200"],
            "from_date": [dateutil.parser.isoparse("2019-01-01T00:00:00.000Z")],
            "to_date": [dateutil.parser.isoparse("2019-12-31T00:00:00.000Z")],
            "value": [1.3],
            "quality": [None],
        }
    )

    df = filter_by_date_and_resolution(df, date="2019-05/2019-09", resolution=Resolution.ANNUAL)
    assert not df.empty

    df = filter_by_date_and_resolution(df, date="2020/2022", resolution=Resolution.ANNUAL)
    assert df.empty

    df = filter_by_date_and_resolution(df, date="2020", resolution=Resolution.ANNUAL)
    assert df.empty


@pytest.mark.sql
def test_filter_by_sql():
    """Test filter by sql statement"""
    df = ExportMixin(df=df_data).filter_by_sql(
        "SELECT * FROM data WHERE parameter='temperature_air_max_200' AND value < 1.5"
    )
    assert not df.empty

    df = ExportMixin(df=df_data).filter_by_sql(
        "SELECT * FROM data WHERE parameter='temperature_air_max_200' AND value > 1.5"
    )
    assert df.empty


def test_format_json():
    """Test export of DataFrame to json"""
    output = ExportMixin(df=df_data).to_json()

    response = json.loads(output)
    station_ids = {reading["station_id"] for reading in response}

    assert "01048" in station_ids


def test_format_geojson():
    """Test export of DataFrame to geojson"""
    output = StationsResult(df=df_station, stations=None).to_geojson()

    response = json.loads(output)

    station_names = {station["properties"]["name"] for station in response["features"]}

    assert "Freyung vorm Wald" in station_names


def test_format_csv():
    """Test export of DataFrame to csv"""
    output = ExportMixin(df=df_data).to_csv().strip()

    assert "station_id,dataset,parameter,date,value,quality" in output
    assert "01048,climate_summary,temperature_air_max_200,2019-12-28T00-00-00,1.3," in output


@pytest.mark.remote
def test_request():
    """Test general data request"""
    Settings.tidy = True
    Settings.humanize = True
    Settings.si_units = True

    request = DwdObservationRequest(
        parameter=DwdObservationDataset.CLIMATE_SUMMARY,
        resolution=DwdObservationResolution.DAILY,
        period=DwdObservationPeriod.RECENT,
    ).filter_by_station_id(station_id=[1048])

    df = request.values.all().df

    assert not df.empty


def test_export_unknown():
    """Test export of DataFrame to unknown format"""
    Settings.tidy = True
    Settings.humanize = True
    Settings.si_units = True

    request = DwdObservationRequest(
        parameter=DwdObservationDataset.CLIMATE_SUMMARY,
        resolution=DwdObservationResolution.DAILY,
        period=DwdObservationPeriod.RECENT,
    ).filter_by_station_id(
        station_id=[1048],
    )

    df = request.values.all().df

    with pytest.raises(KeyError) as ex:
        ExportMixin(df=df).to_target("file:///test.foobar")

    ex.match("Unknown export file type")


def test_export_spreadsheet(tmpdir_factory):
    """Test export of DataFrame to spreadsheet"""
    Settings.tidy = False
    Settings.humanize = True
    Settings.si_units = False

    # 1. Request data and save to .xlsx file.
    request = DwdObservationRequest(
        parameter=DwdObservationDataset.CLIMATE_SUMMARY,
        resolution=DwdObservationResolution.DAILY,
        start_date="2019",
        end_date="2020",
    ).filter_by_station_id(
        station_id=[1048],
    )

    df = request.values.all().df

    filename = tmpdir_factory.mktemp("data").join("observations.xlsx")
    ExportMixin(df=df).to_target(f"file://{filename}")

    workbook = openpyxl.load_workbook(filename=filename)
    worksheet = workbook.active

    # 2. Validate some details of .xlsx file.

    # Validate header row.
    header = list(worksheet.iter_cols(min_row=1, max_row=1, values_only=True))
    assert header == [
        ("station_id",),
        ("dataset",),
        ("date",),
        ("qn_3",),
        ("wind_gust_max",),
        ("wind_speed",),
        ("qn_4",),
        ("precipitation_height",),
        ("precipitation_form",),
        ("sunshine_duration",),
        ("snow_depth",),
        ("cloud_cover_total",),
        ("pressure_vapor",),
        ("pressure_air_site",),
        ("temperature_air_mean_200",),
        ("humidity",),
        ("temperature_air_max_200",),
        ("temperature_air_min_200",),
        ("temperature_air_min_005",),
    ]

    # Validate number of records.
    assert worksheet.max_row == 367

    first_record = list(worksheet.iter_cols(min_row=2, max_row=2, values_only=True))
    assert first_record == [
        ("01048",),
        ("climate_summary",),
        ("2019-01-01T00:00:00+00:00",),
        (10,),
        (19.9,),
        (8.5,),
        (10,),
        (0.9,),
        (8,),
        (0,),
        (0,),
        (7.4,),
        (7.9,),
        (991.9,),
        (5.9,),
        (84,),
        (7.5,),
        (2,),
        (1.5,),
    ]

    last_record = list(worksheet.iter_cols(min_row=worksheet.max_row, max_row=worksheet.max_row, values_only=True))
    assert last_record == [
        ("01048",),
        ("climate_summary",),
        ("2020-01-01T00:00:00+00:00",),
        (10,),
        (6.9,),
        (3.2,),
        (3,),
        (0,),
        (0,),
        (3.933,),
        (0,),
        (4.2,),
        (5.7,),
        (1005.11,),
        (2.4,),
        (79,),
        (5.6,),
        (-2.8,),
        (-4.6,),
    ]

    os.unlink(filename)


@pytest.mark.remote
def test_export_parquet(tmpdir_factory):
    """Test export of DataFrame to parquet"""
    Settings.tidy = False
    Settings.humanize = True
    Settings.si_units = False

    # Request data.
    request = DwdObservationRequest(
        parameter=DwdObservationDataset.CLIMATE_SUMMARY,
        resolution=DwdObservationResolution.DAILY,
        start_date="2019",
        end_date="2020",
    ).filter_by_station_id(
        station_id=[1048],
    )

    df = request.values.all().df

    # Save to Parquet file.
    filename = tmpdir_factory.mktemp("data").join("observations.parquet")
    ExportMixin(df=df).to_target(f"file://{filename}")

    # Read back Parquet file.
    table = pq.read_table(filename)

    # Validate dimensions.
    assert table.num_columns == 19
    assert table.num_rows == 366

    # Validate column names.
    assert table.column_names == [
        "station_id",
        "dataset",
        "date",
        "qn_3",
        "wind_gust_max",
        "wind_speed",
        "qn_4",
        "precipitation_height",
        "precipitation_form",
        "sunshine_duration",
        "snow_depth",
        "cloud_cover_total",
        "pressure_vapor",
        "pressure_air_site",
        "temperature_air_mean_200",
        "humidity",
        "temperature_air_max_200",
        "temperature_air_min_200",
        "temperature_air_min_005",
    ]

    # Validate content.
    data = table.to_pydict()

    assert data["date"][0] == datetime.datetime(2019, 1, 1, 0, 0, tzinfo=datetime.timezone.utc)
    assert data["temperature_air_min_005"][0] == 1.5
    assert data["date"][-1] == datetime.datetime(2020, 1, 1, 0, 0, tzinfo=datetime.timezone.utc)
    assert data["temperature_air_min_005"][-1] == -4.6

    os.unlink(filename)


@pytest.mark.remote
def test_export_zarr(tmpdir_factory):
    """Test export of DataFrame to zarr"""
    Settings.tidy = False
    Settings.humanize = True
    Settings.si_units = False

    # Request data.
    request = DwdObservationRequest(
        parameter=DwdObservationDataset.CLIMATE_SUMMARY,
        resolution=DwdObservationResolution.DAILY,
        start_date="2019",
        end_date="2020",
    ).filter_by_station_id(
        station_id=[1048],
    )

    df = request.values.all().df

    # Save to Zarr group.
    filename = tmpdir_factory.mktemp("data").join("observations.zarr")
    ExportMixin(df=df).to_target(f"file://{filename}")

    # Read back Zarr group.
    group = zarr.open(str(filename), mode="r")

    # Validate dimensions.
    assert len(group) == 20
    assert len(group.index) == 366

    # Validate column names.
    assert set(group.keys()) == {
        "index",
        "station_id",
        "dataset",
        "date",
        "qn_3",
        "wind_gust_max",
        "wind_speed",
        "qn_4",
        "precipitation_height",
        "precipitation_form",
        "sunshine_duration",
        "snow_depth",
        "cloud_cover_total",
        "pressure_vapor",
        "pressure_air_site",
        "temperature_air_mean_200",
        "humidity",
        "temperature_air_max_200",
        "temperature_air_min_200",
        "temperature_air_min_005",
    }

    # Validate content.
    data = group

    assert data["date"][0] == np.datetime64(datetime.datetime(2019, 1, 1, 0, 0, tzinfo=datetime.timezone.utc))
    assert data["temperature_air_min_005"][0] == 1.5
    assert data["date"][-1] == np.datetime64(datetime.datetime(2020, 1, 1, 0, 0, tzinfo=datetime.timezone.utc))
    assert data["temperature_air_min_005"][-1] == -4.6

    shutil.rmtree(filename)


@pytest.mark.remote
def test_export_feather(tmpdir_factory):
    """Test export of DataFrame to feather"""
    Settings.tidy = False
    Settings.humanize = True
    Settings.si_units = False

    # Request data
    request = DwdObservationRequest(
        parameter=DwdObservationDataset.CLIMATE_SUMMARY,
        resolution=DwdObservationResolution.DAILY,
        start_date="2019",
        end_date="2020",
    ).filter_by_station_id(
        station_id=[1048],
    )

    df = request.values.all().df

    # Save to Feather file.
    filename = tmpdir_factory.mktemp("data").join("observations.feather")
    ExportMixin(df=df).to_target(f"file://{filename}")

    # Read back Feather file.
    table = feather.read_table(filename)

    # Validate dimensions.
    assert table.num_columns == 19
    assert table.num_rows == 366

    # Validate column names.
    assert table.column_names == [
        "station_id",
        "dataset",
        "date",
        "qn_3",
        "wind_gust_max",
        "wind_speed",
        "qn_4",
        "precipitation_height",
        "precipitation_form",
        "sunshine_duration",
        "snow_depth",
        "cloud_cover_total",
        "pressure_vapor",
        "pressure_air_site",
        "temperature_air_mean_200",
        "humidity",
        "temperature_air_max_200",
        "temperature_air_min_200",
        "temperature_air_min_005",
    ]

    # Validate content.
    data = table.to_pydict()

    assert data["date"][0] == datetime.datetime(2019, 1, 1, 0, 0, tzinfo=datetime.timezone.utc)
    assert data["temperature_air_min_005"][0] == 1.5
    assert data["date"][-1] == datetime.datetime(2020, 1, 1, 0, 0, tzinfo=datetime.timezone.utc)
    assert data["temperature_air_min_005"][-1] == -4.6

    os.unlink(filename)


@pytest.mark.remote
def test_export_sqlite(tmpdir_factory):
    """Test export of DataFrame to sqlite db"""
    Settings.tidy = False
    Settings.humanize = True
    Settings.si_units = False

    request = DwdObservationRequest(
        parameter=DwdObservationDataset.CLIMATE_SUMMARY,
        resolution=DwdObservationResolution.DAILY,
        start_date="2019",
        end_date="2020",
    ).filter_by_station_id(
        station_id=[1048],
    )

    filename = tmpdir_factory.mktemp("data").join("observations.sqlite")

    df = request.values.all().df
    ExportMixin(df=df).to_target(f"sqlite:///{filename}?table=testdrive")

    connection = sqlite3.connect(filename)
    cursor = connection.cursor()
    cursor.execute("SELECT * FROM testdrive")
    results = cursor.fetchall()
    cursor.close()
    connection.close()

    assert results[0] == (
        "01048",
        "climate_summary",
        "2019-01-01 00:00:00.000000",
        10.0,
        19.9,
        8.5,
        10.0,
        0.9,
        8.0,
        0.0,
        0.0,
        7.4,
        7.9,
        991.9,
        5.9,
        84.0,
        7.5,
        2.0,
        1.5,
    )

    assert results[-1] == (
        "01048",
        "climate_summary",
        "2020-01-01 00:00:00.000000",
        10,
        6.9,
        3.2,
        3,
        0.0,
        0,
        3.933,
        0,
        4.2,
        5.7,
        1005.11,
        2.4,
        79.0,
        5.6,
        -2.8,
        -4.6,
    )


@pytest.mark.remote
def test_export_cratedb():
    """Test export of DataFrame to cratedb"""
    Settings.tidy = True
    Settings.humanize = True
    Settings.si_units = False

    request = DwdObservationRequest(
        parameter=DwdObservationDataset.CLIMATE_SUMMARY,
        resolution=DwdObservationResolution.DAILY,
        period=DwdObservationPeriod.RECENT,
    ).filter_by_station_id(
        station_id=[1048],
    )

    with mock.patch(
        "pandas.DataFrame.to_sql",
    ) as mock_to_sql:

        df = request.values.all().df
        ExportMixin(df=df).to_target("crate://localhost/?database=test&table=testdrive")

        mock_to_sql.assert_called_once_with(
            name="testdrive",
            con="crate://localhost",
            schema="test",
            if_exists="replace",
            index=False,
            chunksize=5000,
        )


@surrogate("duckdb.connect")
@pytest.mark.remote
def test_export_duckdb():
    """Test export of DataFrame to duckdb"""
    Settings.tidy = True
    Settings.humanize = True
    Settings.si_units = False

    request = DwdObservationRequest(
        parameter=DwdObservationDataset.CLIMATE_SUMMARY,
        resolution=DwdObservationResolution.DAILY,
        period=DwdObservationPeriod.RECENT,
    ).filter_by_station_id(station_id=[1048])

    mock_connection = mock.MagicMock()
    with mock.patch("duckdb.connect", side_effect=[mock_connection], create=True) as mock_connect:

        df = request.values.all().df
        ExportMixin(df=df).to_target("duckdb:///test.duckdb?table=testdrive")

        mock_connect.assert_called_once_with(database="test.duckdb", read_only=False)
        mock_connection.register.assert_called_once()
        mock_connection.execute.assert_called()
        mock_connection.table.assert_called_once_with("testdrive")
        mock_connection.close.assert_called_once()


@surrogate("influxdb.InfluxDBClient")
@pytest.mark.remote
def test_export_influxdb1_tabular():
    """Test export of DataFrame to influxdb v1"""
    Settings.tidy = False
    Settings.humanize = True
    Settings.si_units = False

    request = DwdObservationRequest(
        parameter=DwdObservationDataset.CLIMATE_SUMMARY,
        resolution=DwdObservationResolution.DAILY,
        period=DwdObservationPeriod.RECENT,
    ).filter_by_station_id(station_id=[1048])

    mock_client = mock.MagicMock()
    with mock.patch(
        "influxdb.InfluxDBClient",
        side_effect=[mock_client],
        create=True,
    ) as mock_connect:

        df = request.values.all().df
        ExportMixin(df=df).to_target("influxdb://localhost/?database=dwd&table=weather")

        mock_connect.assert_called_once_with(
            host="localhost",
            port=8086,
            username=None,
            password=None,
            database="dwd",
            ssl=False,
        )
        mock_client.create_database.assert_called_once_with("dwd")
        mock_client.write_points.assert_called_once()

        mock_client.write_points.assert_called_with(
            points=mock.ANY,
            batch_size=50000,
        )

        points = mock_client.write_points.call_args.kwargs["points"]
        assert points[0]["measurement"] == "weather"
        assert list(points[0]["tags"].keys()) == [
            "station_id",
            "qn_3",
            "qn_4",
            "dataset",
        ]
        assert list(points[0]["fields"].keys()) == [
            "wind_gust_max",
            "wind_speed",
            "precipitation_height",
            "precipitation_form",
            "sunshine_duration",
            "snow_depth",
            "cloud_cover_total",
            "pressure_vapor",
            "pressure_air_site",
            "temperature_air_mean_200",
            "humidity",
            "temperature_air_max_200",
            "temperature_air_min_200",
            "temperature_air_min_005",
        ]


@surrogate("influxdb.InfluxDBClient")
@pytest.mark.remote
def test_export_influxdb1_tidy():
    """Test export of DataFrame to influxdb v1"""
    Settings.tidy = True
    Settings.humanize = True
    Settings.si_units = False

    request = DwdObservationRequest(
        parameter=DwdObservationDataset.CLIMATE_SUMMARY,
        resolution=DwdObservationResolution.DAILY,
        period=DwdObservationPeriod.RECENT,
    ).filter_by_station_id(station_id=[1048])

    mock_client = mock.MagicMock()
    with mock.patch(
        "influxdb.InfluxDBClient",
        side_effect=[mock_client],
        create=True,
    ) as mock_connect:

        df = request.values.all().df
        ExportMixin(df=df).to_target("influxdb://localhost/?database=dwd&table=weather")

        mock_connect.assert_called_once_with(
            host="localhost",
            port=8086,
            username=None,
            password=None,
            database="dwd",
            ssl=False,
        )
        mock_client.create_database.assert_called_once_with("dwd")
        mock_client.write_points.assert_called_once()

        mock_client.write_points.assert_called_with(
            points=mock.ANY,
            batch_size=50000,
        )

        points = mock_client.write_points.call_args.kwargs["points"]
        assert points[0]["measurement"] == "weather"
        assert list(points[0]["tags"].keys()) == [
            "station_id",
            "quality",
            "dataset",
            "parameter",
        ]
        assert list(points[0]["fields"].keys()) == [
            "value",
        ]


@surrogate("influxdb_client.InfluxDBClient")
@surrogate("influxdb_client.Point")
@surrogate("influxdb_client.client.write_api.SYNCHRONOUS")
@pytest.mark.remote
def test_export_influxdb2_tabular():
    """Test export of DataFrame to influxdb v2"""
    Settings.tidy = True
    Settings.humanize = True
    Settings.si_units = False

    request = DwdObservationRequest(
        parameter=DwdObservationDataset.CLIMATE_SUMMARY,
        resolution=DwdObservationResolution.DAILY,
        period=DwdObservationPeriod.RECENT,
    ).filter_by_station_id(station_id=[1048])

    mock_client = mock.MagicMock()
    with mock.patch(
        "influxdb_client.InfluxDBClient",
        side_effect=[mock_client],
        create=True,
    ) as mock_connect:

        with mock.patch(
            "influxdb_client.Point",
            create=True,
        ):

            df = request.values.all().df
            ExportMixin(df=df).to_target("influxdb2://orga:token@localhost/?database=dwd&table=weather")

            mock_connect.assert_called_once_with(url="http://localhost:8086", org="orga", token="token")


@surrogate("influxdb_client.InfluxDBClient")
@surrogate("influxdb_client.Point")
@surrogate("influxdb_client.client.write_api.SYNCHRONOUS")
@pytest.mark.remote
def test_export_influxdb2_tidy():
    """Test export of DataFrame to influxdb v2"""
    Settings.tidy = True
    Settings.humanize = True
    Settings.si_units = False

    request = DwdObservationRequest(
        parameter=DwdObservationDataset.CLIMATE_SUMMARY,
        resolution=DwdObservationResolution.DAILY,
        period=DwdObservationPeriod.RECENT,
    ).filter_by_station_id(station_id=[1048])

    mock_client = mock.MagicMock()
    with mock.patch(
        "influxdb_client.InfluxDBClient",
        side_effect=[mock_client],
        create=True,
    ) as mock_connect:

        with mock.patch(
            "influxdb_client.Point",
            create=True,
        ):

            df = request.values.all().df
            ExportMixin(df=df).to_target("influxdb2://orga:token@localhost/?database=dwd&table=weather")

            mock_connect.assert_called_once_with(url="http://localhost:8086", org="orga", token="token")
