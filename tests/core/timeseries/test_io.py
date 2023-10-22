# -*- coding: utf-8 -*-
# Copyright (C) 2018-2021, earthobservations developers.
# Distributed under the MIT License. See LICENSE for more info.
import datetime as dt
import json
import sqlite3
from unittest import mock

import polars as pl
import pytest
from surrogate import surrogate
from zoneinfo import ZoneInfo

from wetterdienst import Provider
from wetterdienst.core.process import filter_by_date
from wetterdienst.core.timeseries.export import ExportMixin
from wetterdienst.core.timeseries.result import (
    InterpolatedValuesResult,
    StationsFilter,
    StationsResult,
    SummarizedValuesResult,
    ValuesResult,
)
from wetterdienst.provider.dwd.observation import (
    DwdObservationDataset,
    DwdObservationPeriod,
    DwdObservationRequest,
    DwdObservationResolution,
)


@pytest.fixture
def dwd_climate_summary_tabular_columns():
    return [
        "station_id",
        "dataset",
        "date",
        "wind_gust_max",
        "qn_wind_gust_max",
        "wind_speed",
        "qn_wind_speed",
        "precipitation_height",
        "qn_precipitation_height",
        "precipitation_form",
        "qn_precipitation_form",
        "sunshine_duration",
        "qn_sunshine_duration",
        "snow_depth",
        "qn_snow_depth",
        "cloud_cover_total",
        "qn_cloud_cover_total",
        "pressure_vapor",
        "qn_pressure_vapor",
        "pressure_air_site",
        "qn_pressure_air_site",
        "temperature_air_mean_200",
        "qn_temperature_air_mean_200",
        "humidity",
        "qn_humidity",
        "temperature_air_max_200",
        "qn_temperature_air_max_200",
        "temperature_air_min_200",
        "qn_temperature_air_min_200",
        "temperature_air_min_005",
        "qn_temperature_air_min_005",
    ]


@pytest.fixture
def dwd_metadata():
    return {
        "producer": {
            "doi": "10.5281/zenodo.3960624",
            "name": "Wetterdienst",
            "url": "https://github.com/earthobservations/wetterdienst",
        },
        "provider": {
            "copyright": "© Deutscher Wetterdienst (DWD), Climate Data Center (CDC)",
            "country": "Germany",
            "name_english": "German Weather Service",
            "name_local": "Deutscher Wetterdienst",
            "url": "https://opendata.dwd.de/climate_environment/CDC/",
        },
    }


@pytest.fixture
def df_stations():
    return pl.DataFrame(
        {
            "station_id": ["01048"],
            "from_date": [dt.datetime(1957, 5, 1, tzinfo=ZoneInfo("UTC"))],
            "to_date": [dt.datetime(1995, 11, 30, tzinfo=ZoneInfo("UTC"))],
            "height": [645.0],
            "latitude": [48.8049],
            "longitude": [13.5528],
            "name": ["Freyung vorm Wald"],
            "state": ["Bayern"],
        }
    )


@pytest.fixture
def stations_mock():
    class StationsMock:
        _provider = Provider.DWD

    return StationsMock


@pytest.fixture
def stations_result_mock(df_stations, stations_mock):
    return StationsResult(
        df=df_stations,
        df_all=df_stations,
        stations_filter=StationsFilter.ALL,
        stations=stations_mock,
    )


@pytest.fixture
def df_values():
    return pl.DataFrame(
        [
            {
                "station_id": "01048",
                "dataset": "climate_summary",
                "parameter": "temperature_air_max_200",
                "date": dt.datetime(2019, 1, 1, tzinfo=ZoneInfo("UTC")),
                "value": 1.3,
                "quality": None,
            },
            {
                "station_id": "01048",
                "dataset": "climate_summary",
                "parameter": "temperature_air_max_200",
                "date": dt.datetime(2019, 12, 1, tzinfo=ZoneInfo("UTC")),
                "value": 1.0,
                "quality": None,
            },
            {
                "station_id": "01048",
                "dataset": "climate_summary",
                "parameter": "temperature_air_max_200",
                "date": dt.datetime(2019, 12, 28, tzinfo=ZoneInfo("UTC")),
                "value": 1.3,
                "quality": None,
            },
            {
                "station_id": "01048",
                "dataset": "climate_summary",
                "parameter": "temperature_air_max_200",
                "date": dt.datetime(2020, 1, 1, tzinfo=ZoneInfo("UTC")),
                "value": 2.0,
                "quality": None,
            },
            {
                "station_id": "01048",
                "dataset": "climate_summary",
                "parameter": "temperature_air_max_200",
                "date": dt.datetime(2021, 1, 1, tzinfo=ZoneInfo("UTC")),
                "value": 3.0,
                "quality": None,
            },
            {
                "station_id": "01048",
                "dataset": "climate_summary",
                "parameter": "temperature_air_max_200",
                "date": dt.datetime(2022, 1, 1, tzinfo=ZoneInfo("UTC")),
                "value": 4.0,
                "quality": None,
            },
        ],
        schema={
            "station_id": pl.Utf8,
            "dataset": pl.Utf8,
            "parameter": pl.Utf8,
            "date": pl.Datetime(time_zone="UTC"),
            "value": pl.Float64,
            "quality": pl.Float64,
        },
    )


@pytest.fixture
def df_interpolated_values():
    return pl.DataFrame(
        [
            {
                "date": dt.datetime(2019, 1, 1, tzinfo=ZoneInfo("UTC")),
                "parameter": "temperature_air_max_200",
                "value": 1.3,
                "distance_mean": 0.0,
                "station_ids": ["01048"],
            },
        ],
        schema={
            "date": pl.Datetime(time_zone="UTC"),
            "parameter": pl.Utf8,
            "value": pl.Float64,
            "distance_mean": pl.Float64,
            "station_ids": pl.List(pl.Utf8),
        },
    )


@pytest.fixture
def df_summarized_values():
    return pl.DataFrame(
        [
            {
                "date": dt.datetime(2019, 1, 1, tzinfo=ZoneInfo("UTC")),
                "parameter": "temperature_air_max_200",
                "value": 1.3,
                "distance": 0.0,
                "station_id": "01048",
            },
        ],
        schema={
            "date": pl.Datetime(time_zone="UTC"),
            "parameter": pl.Utf8,
            "value": pl.Float64,
            "distance": pl.Float64,
            "station_id": pl.Utf8,
        },
    )


def test_stations_to_dict(df_stations):
    data = StationsResult(
        df=df_stations,
        df_all=df_stations,
        stations_filter=StationsFilter.ALL,
        stations=None,
    ).to_dict()
    assert data.keys() == {"stations"}
    assert data["stations"] == [
        {
            "station_id": "01048",
            "from_date": "1957-05-01T00:00:00+00:00",
            "to_date": "1995-11-30T00:00:00+00:00",
            "height": 645.0,
            "latitude": 48.8049,
            "longitude": 13.5528,
            "name": "Freyung vorm Wald",
            "state": "Bayern",
        },
    ]


def test_stations_to_dict_with_metadata(df_stations, stations_mock, dwd_metadata):
    data = StationsResult(
        df=df_stations,
        df_all=df_stations,
        stations_filter=StationsFilter.ALL,
        stations=stations_mock,
    ).to_dict(with_metadata=True)
    assert data.keys() == {"stations", "metadata"}
    assert data["metadata"] == dwd_metadata


def test_stations_to_ogc_feature_collection(df_stations):
    data = StationsResult(
        df=df_stations,
        df_all=df_stations,
        stations_filter=StationsFilter.ALL,
        stations=None,
    ).to_ogc_feature_collection()
    assert data.keys() == {"data"}
    assert data["data"]["features"][0] == {
        "geometry": {"coordinates": [13.5528, 48.8049, 645.0], "type": "Point"},
        "properties": {
            "from_date": "1957-05-01T00:00:00+00:00",
            "id": "01048",
            "name": "Freyung vorm Wald",
            "state": "Bayern",
            "to_date": "1995-11-30T00:00:00+00:00",
        },
        "type": "Feature",
    }


def test_stations_to_ogc_feature_collection_with_metadata(df_stations, stations_mock, dwd_metadata):
    data = StationsResult(
        df=df_stations,
        df_all=df_stations,
        stations_filter=StationsFilter.ALL,
        stations=stations_mock,
    ).to_ogc_feature_collection(with_metadata=True)
    assert data.keys() == {"data", "metadata"}
    assert data["metadata"] == dwd_metadata


def test_stations_format_json(df_stations):
    """Test export of DataFrame to json"""
    output = StationsResult(
        df=df_stations,
        df_all=df_stations,
        stations_filter=StationsFilter.ALL,
        stations=None,
    ).to_json()
    response = json.loads(output)
    assert response.keys() == {"stations"}
    station_ids = {station["station_id"] for station in response["stations"]}
    assert "01048" in station_ids


def test_stations_format_geojson(df_stations, stations_mock):
    """Test export of DataFrame to geojson"""
    output = StationsResult(
        df=df_stations,
        df_all=df_stations,
        stations_filter=StationsFilter.ALL,
        stations=stations_mock,
    ).to_geojson()
    response = json.loads(output)
    assert response.keys() == {"data"}
    station_names = {station["properties"]["name"] for station in response["data"]["features"]}
    assert "Freyung vorm Wald" in station_names


def test_stations_format_csv(df_stations):
    """Test export of DataFrame to csv"""
    output = (
        StationsResult(
            df=df_stations,
            df_all=df_stations,
            stations_filter=StationsFilter.ALL,
            stations=None,
        )
        .to_csv()
        .strip()
    )
    assert "station_id,from_date,to_date,height,latitude,longitude,name,state" in output
    assert (
        "01048,1957-05-01T00:00:00+00:00,1995-11-30T00:00:00+00:00,645.0,48.8049,13.5528,Freyung vorm Wald,Bayern"
        in output
    )


def test_values_to_dict(df_values):
    data = ValuesResult(stations=None, values=None, df=df_values[0, :]).to_dict()
    assert data.keys() == {"values"}
    assert data["values"] == [
        {
            "dataset": "climate_summary",
            "date": "2019-01-01T00:00:00+00:00",
            "parameter": "temperature_air_max_200",
            "quality": None,
            "station_id": "01048",
            "value": 1.3,
        },
    ]


def test_values_to_dict_with_metadata(df_values, stations_result_mock, dwd_metadata):
    data = ValuesResult(stations=stations_result_mock, values=None, df=df_values[0, :]).to_dict(with_metadata=True)
    assert data.keys() == {"values", "metadata"}
    assert data["metadata"] == dwd_metadata


def test_values_to_ogc_feature_collection(df_values, stations_result_mock):
    data = ValuesResult(stations=stations_result_mock, values=None, df=df_values[0, :]).to_ogc_feature_collection()
    assert data.keys() == {"data"}
    assert data["data"]["features"][0] == {
        "geometry": {"coordinates": [13.5528, 48.8049, 645.0], "type": "Point"},
        "properties": {
            "id": "01048",
            "name": "Freyung vorm Wald",
            "from_date": "1957-05-01T00:00:00+00:00",
            "to_date": "1995-11-30T00:00:00+00:00",
            "state": "Bayern",
        },
        "type": "Feature",
        "values": [
            {
                "dataset": "climate_summary",
                "date": "2019-01-01T00:00:00+00:00",
                "parameter": "temperature_air_max_200",
                "quality": None,
                "value": 1.3,
            }
        ],
    }


def test_values_to_ogc_feature_collection_with_metadata(df_values, stations_result_mock, dwd_metadata):
    data = ValuesResult(stations=stations_result_mock, values=None, df=df_values[0, :]).to_ogc_feature_collection(
        with_metadata=True
    )
    assert data.keys() == {"data", "metadata"}
    assert data["metadata"] == dwd_metadata


def test_values_format_json(df_values):
    """Test export of DataFrame to json"""
    output = ValuesResult(stations=None, values=None, df=df_values).to_json()
    response = json.loads(output)
    assert response.keys() == {"values"}
    station_ids = {reading["station_id"] for reading in response["values"]}
    assert "01048" in station_ids


def test_values_format_geojson(df_values, stations_result_mock):
    """Test export of DataFrame to geojson"""
    output = ValuesResult(df=df_values, stations=stations_result_mock, values=None).to_geojson()
    response = json.loads(output)
    assert response.keys() == {"data"}
    item = response["data"]["features"][0]["values"][0]
    assert item == {
        "dataset": "climate_summary",
        "parameter": "temperature_air_max_200",
        "date": "2019-01-01T00:00:00+00:00",
        "value": 1.3,
        "quality": None,
    }


def test_values_format_csv(df_values):
    """Test export of DataFrame to csv"""
    output = ValuesResult(stations=None, values=None, df=df_values).to_csv().strip()
    assert "station_id,dataset,parameter,date,value,quality" in output
    assert "01048,climate_summary,temperature_air_max_200,2019-12-28T00:00:00+00:00,1.3," in output


def test_interpolated_values_to_dict(df_interpolated_values):
    data = InterpolatedValuesResult(stations=None, df=df_interpolated_values, latlon=(1, 2)).to_dict()
    assert data.keys() == {"values"}
    assert data["values"] == [
        {
            "date": "2019-01-01T00:00:00+00:00",
            "parameter": "temperature_air_max_200",
            "value": 1.3,
            "distance_mean": 0.0,
            "station_ids": ["01048"],
        }
    ]


def test_interpolated_values_to_dict_with_metadata(df_interpolated_values, stations_result_mock, dwd_metadata):
    data = InterpolatedValuesResult(stations=stations_result_mock, df=df_interpolated_values, latlon=(1, 2)).to_dict(
        with_metadata=True
    )
    assert data.keys() == {"values", "metadata"}
    assert data["metadata"] == dwd_metadata


def test_interpolated_values_to_ogc_feature_collection(df_interpolated_values, stations_result_mock):
    data = InterpolatedValuesResult(
        stations=stations_result_mock, df=df_interpolated_values, latlon=(1.2345, 2.3456)
    ).to_ogc_feature_collection()
    assert data.keys() == {"data"}
    assert data["data"]["features"][0] == {
        "geometry": {"coordinates": [2.3456, 1.2345], "type": "Point"},
        "properties": {"name": "interpolation(lat=1.2345,lon=2.3456)"},
        "stations": [
            {
                "station_id": "01048",
                "latitude": 48.8049,
                "longitude": 13.5528,
                "height": 645.0,
                "from_date": "1957-05-01T00:00:00+00:00",
                "to_date": "1995-11-30T00:00:00+00:00",
                "name": "Freyung vorm Wald",
                "state": "Bayern",
            }
        ],
        "type": "Feature",
        "values": [
            {
                "date": "2019-01-01T00:00:00+00:00",
                "parameter": "temperature_air_max_200",
                "value": 1.3,
                "distance_mean": 0.0,
                "station_ids": ["01048"],
            }
        ],
    }


def test_interpolated_values_to_ogc_feature_collection_with_metadata(
    df_interpolated_values, stations_result_mock, dwd_metadata
):
    data = InterpolatedValuesResult(
        stations=stations_result_mock, df=df_interpolated_values, latlon=(1.2345, 2.3456)
    ).to_ogc_feature_collection(with_metadata=True)
    assert data.keys() == {"data", "metadata"}
    assert data["metadata"] == dwd_metadata


def test_summarized_values_to_dict(df_summarized_values):
    data = SummarizedValuesResult(stations=None, df=df_summarized_values, latlon=(1.2345, 2.3456)).to_dict()
    assert data.keys() == {"values"}
    assert data["values"] == [
        {
            "date": "2019-01-01T00:00:00+00:00",
            "parameter": "temperature_air_max_200",
            "value": 1.3,
            "distance": 0.0,
            "station_id": "01048",
        }
    ]


def test_summarized_values_to_dict_with_metadata(df_summarized_values, stations_result_mock, dwd_metadata):
    data = SummarizedValuesResult(
        stations=stations_result_mock, df=df_summarized_values, latlon=(1.2345, 2.3456)
    ).to_dict(with_metadata=True)
    assert data.keys() == {"values", "metadata"}
    assert data["metadata"] == dwd_metadata


def test_summarized_values_to_ogc_feature_collection(df_summarized_values, stations_result_mock):
    data = SummarizedValuesResult(
        stations=stations_result_mock, df=df_summarized_values, latlon=(1.2345, 2.3456)
    ).to_ogc_feature_collection()
    assert data.keys() == {"data"}
    assert data["data"]["features"][0] == {
        "geometry": {"coordinates": [2.3456, 1.2345], "type": "Point"},
        "properties": {"name": "summary(lat=1.2345,lon=2.3456)"},
        "stations": [
            {
                "station_id": "01048",
                "latitude": 48.8049,
                "longitude": 13.5528,
                "height": 645.0,
                "from_date": "1957-05-01T00:00:00+00:00",
                "to_date": "1995-11-30T00:00:00+00:00",
                "name": "Freyung vorm Wald",
                "state": "Bayern",
            }
        ],
        "type": "Feature",
        "values": [
            {
                "date": "2019-01-01T00:00:00+00:00",
                "parameter": "temperature_air_max_200",
                "value": 1.3,
                "distance": 0.0,
                "station_id": "01048",
            }
        ],
    }


def test_summarized_values_to_ogc_feature_collection_with_metadata(
    df_summarized_values, stations_result_mock, dwd_metadata
):
    data = SummarizedValuesResult(
        stations=stations_result_mock, df=df_summarized_values, latlon=(1.2345, 2.3456)
    ).to_ogc_feature_collection(with_metadata=True)
    assert data.keys() == {"data", "metadata"}
    assert data["metadata"] == dwd_metadata


def test_filter_by_date(df_values):
    """Test filter by date"""
    df = filter_by_date(df_values, "2019-12-28")
    assert not df.is_empty()
    df = filter_by_date(df_values, "2019-12-27")
    assert df.is_empty()


def test_filter_by_date_interval(df_values):
    """Test filter by date interval"""
    df = filter_by_date(df_values, "2019-12-27/2019-12-29")
    assert not df.is_empty()
    df = filter_by_date(df_values, "2019-12/2020-01")
    assert df.get_column("value").to_list() == [1.0, 1.3, 2.0]
    df = filter_by_date(df, date="2020/2022")
    assert not df.is_empty()
    df = filter_by_date(df, date="2020")
    assert not df.is_empty()


@pytest.mark.sql
def test_filter_by_sql(df_values):
    """Test filter by sql statement"""
    df = ExportMixin(df=df_values).filter_by_sql(
        sql="SELECT * FROM data WHERE parameter='temperature_air_max_200' AND value < 1.5"
    )
    assert not df.is_empty()
    df = ExportMixin(df=df_values).filter_by_sql(
        sql="SELECT * FROM data WHERE parameter='temperature_air_max_200' AND value > 4"
    )
    assert df.is_empty()


@pytest.mark.remote
def test_request(default_settings):
    """Test general data request"""
    request = DwdObservationRequest(
        parameter=DwdObservationDataset.CLIMATE_SUMMARY,
        resolution=DwdObservationResolution.DAILY,
        period=DwdObservationPeriod.RECENT,
        settings=default_settings,
    ).filter_by_station_id(station_id=[1048])
    df = request.values.all().df
    assert not df.is_empty()


@pytest.mark.remote
def test_export_unknown(default_settings):
    """Test export of DataFrame to unknown format"""
    request = DwdObservationRequest(
        parameter=DwdObservationDataset.CLIMATE_SUMMARY,
        resolution=DwdObservationResolution.DAILY,
        period=DwdObservationPeriod.RECENT,
        settings=default_settings,
    ).filter_by_station_id(
        station_id=[1048],
    )
    df = request.values.all().df
    with pytest.raises(KeyError) as exec_info:
        ExportMixin(df=df).to_target("file:///test.foobar")
    assert exec_info.match("Unknown export file type")


@pytest.mark.remote
def test_export_spreadsheet(tmp_path, settings_si_false_wide_shape):
    """Test export of DataFrame to spreadsheet"""
    import openpyxl

    # 1. Request data and save to .xlsx file.
    request = DwdObservationRequest(
        parameter=DwdObservationDataset.CLIMATE_SUMMARY,
        resolution=DwdObservationResolution.DAILY,
        start_date="2019-01-01",
        end_date="2020-01-01",
        settings=settings_si_false_wide_shape,
    ).filter_by_station_id(
        station_id=[1048],
    )
    df = request.values.all().df
    filename = tmp_path.joinpath("observations.xlsx")
    ExportMixin(df=df).to_target(f"file://{filename}")
    workbook = openpyxl.load_workbook(filename=filename)
    worksheet = workbook.active
    # 2. Validate some details of .xlsx file.
    # Validate header row.
    header = list(worksheet.iter_cols(min_row=1, max_row=1, values_only=True))
    assert header == [
        ("station_id",),
        ("dataset",),
        ("date",),
        ("wind_gust_max",),
        ("qn_wind_gust_max",),
        ("wind_speed",),
        ("qn_wind_speed",),
        ("precipitation_height",),
        ("qn_precipitation_height",),
        ("precipitation_form",),
        ("qn_precipitation_form",),
        ("sunshine_duration",),
        ("qn_sunshine_duration",),
        ("snow_depth",),
        ("qn_snow_depth",),
        ("cloud_cover_total",),
        ("qn_cloud_cover_total",),
        ("pressure_vapor",),
        ("qn_pressure_vapor",),
        ("pressure_air_site",),
        ("qn_pressure_air_site",),
        ("temperature_air_mean_200",),
        ("qn_temperature_air_mean_200",),
        ("humidity",),
        ("qn_humidity",),
        ("temperature_air_max_200",),
        ("qn_temperature_air_max_200",),
        ("temperature_air_min_200",),
        ("qn_temperature_air_min_200",),
        ("temperature_air_min_005",),
        ("qn_temperature_air_min_005",),
    ]
    # Validate number of records.
    assert worksheet.max_row == 367
    first_record = list(worksheet.iter_cols(min_row=2, max_row=2, values_only=True))
    assert first_record == [
        ("01048",),
        ("climate_summary",),
        ("2019-01-01T00:00:00+00:00",),
        (19.9,),
        (10,),
        (8.5,),
        (10,),
        (0.9,),
        (10,),
        (8,),
        (10,),
        (0,),
        (10,),
        (0,),
        (10,),
        (7.4,),
        (10,),
        (7.9,),
        (10,),
        (991.9,),
        (10,),
        (5.9,),
        (10,),
        (84,),
        (10,),
        (7.5,),
        (10,),
        (2,),
        (10,),
        (1.5,),
        (10,),
    ]
    last_record = list(worksheet.iter_cols(min_row=worksheet.max_row, max_row=worksheet.max_row, values_only=True))
    assert last_record == [
        ("01048",),
        ("climate_summary",),
        ("2020-01-01T00:00:00+00:00",),
        (6.9,),
        (10,),
        (3.2,),
        (10,),
        (0,),
        (10,),
        (0,),
        (10,),
        (3.9,),
        (10,),
        (0,),
        (10,),
        (4.2,),
        (10,),
        (5.7,),
        (10,),
        (1005.1,),
        (10,),
        (2.4,),
        (10,),
        (79,),
        (10,),
        (5.6,),
        (10,),
        (-2.8,),
        (10,),
        (-4.6,),
        (10,),
    ]


@pytest.mark.remote
def test_export_parquet(tmp_path, settings_si_false_wide_shape, dwd_climate_summary_tabular_columns):
    """Test export of DataFrame to parquet"""
    pq = pytest.importorskip("pyarrow.parquet")
    # Request data.
    request = DwdObservationRequest(
        parameter=DwdObservationDataset.CLIMATE_SUMMARY,
        resolution=DwdObservationResolution.DAILY,
        start_date="2019-01-01",
        end_date="2020-01-01",
        settings=settings_si_false_wide_shape,
    ).filter_by_station_id(
        station_id=[1048],
    )
    df = request.values.all().df
    # Save to Parquet file.
    filename = tmp_path.joinpath("observation.parquet")
    ExportMixin(df=df).to_target(f"file://{filename}")
    # Read back Parquet file.
    table = pq.read_table(filename)
    # Validate dimensions.
    assert table.num_columns == 31
    assert table.num_rows == 366
    # Validate column names.
    assert table.column_names == dwd_climate_summary_tabular_columns
    # Validate content.
    data = table.to_pydict()
    assert data["date"][0] == dt.datetime(2019, 1, 1, 0, 0, tzinfo=ZoneInfo("UTC"))
    assert data["temperature_air_min_005"][0] == 1.5
    assert data["date"][-1] == dt.datetime(2020, 1, 1, 0, 0, tzinfo=ZoneInfo("UTC"))
    assert data["temperature_air_min_005"][-1] == -4.6


@pytest.mark.remote
def test_export_zarr(tmp_path, settings_si_false_wide_shape, dwd_climate_summary_tabular_columns):
    """Test export of DataFrame to zarr"""
    zarr = pytest.importorskip("zarr")
    # Request data.
    request = DwdObservationRequest(
        parameter=DwdObservationDataset.CLIMATE_SUMMARY,
        resolution=DwdObservationResolution.DAILY,
        start_date="2019-01-01",
        end_date="2020-01-01",
        settings=settings_si_false_wide_shape,
    ).filter_by_station_id(
        station_id=[1048],
    )
    df = request.values.all().df
    # Save to Zarr group.
    filename = tmp_path.joinpath("observation.zarr")
    ExportMixin(df=df).to_target(f"file://{filename}")

    # Read back Zarr group.
    root = zarr.open(filename, mode="r")
    group = root.get("climate_summary")
    # Validate dimensions.
    assert len(group) == 32
    assert len(group.index) == 366
    # Validate column names.
    columns = set(group.keys())
    columns.discard("index")
    assert columns == set(dwd_climate_summary_tabular_columns)
    # Validate content.
    data = group
    assert dt.datetime.fromtimestamp(int(data["date"][0]) / 1e9) == dt.datetime(2019, 1, 1, 0, 0)
    assert data["temperature_air_min_005"][0] == 1.5
    assert dt.datetime.fromtimestamp(int(data["date"][-1]) / 1e9) == dt.datetime(2020, 1, 1, 0, 0)
    assert data["temperature_air_min_005"][-1] == -4.6


@pytest.mark.remote
def test_export_feather(tmp_path, settings_si_false_wide_shape, dwd_climate_summary_tabular_columns):
    """Test export of DataFrame to feather"""
    feather = pytest.importorskip("pyarrow.feather")
    # Request data
    request = DwdObservationRequest(
        parameter=DwdObservationDataset.CLIMATE_SUMMARY,
        resolution=DwdObservationResolution.DAILY,
        start_date="2019-01-01",
        end_date="2020-01-01",
        settings=settings_si_false_wide_shape,
    ).filter_by_station_id(
        station_id=[1048],
    )
    df = request.values.all().df
    # Save to Feather file.
    filename = tmp_path.joinpath("observation.feather")
    ExportMixin(df=df).to_target(f"file://{filename}")
    # Read back Feather file.
    table = feather.read_table(filename)
    # Validate dimensions.
    assert table.num_columns == 31
    assert table.num_rows == 366
    # Validate column names.
    assert table.column_names == dwd_climate_summary_tabular_columns
    # Validate content.
    data = table.to_pydict()
    assert data["date"][0] == dt.datetime(2019, 1, 1, 0, 0, tzinfo=ZoneInfo("UTC"))
    assert data["temperature_air_min_005"][0] == 1.5
    assert data["date"][-1] == dt.datetime(2020, 1, 1, 0, 0, tzinfo=ZoneInfo("UTC"))
    assert data["temperature_air_min_005"][-1] == -4.6


@pytest.mark.remote
def test_export_sqlite(tmp_path, settings_si_false_wide_shape):
    """Test export of DataFrame to sqlite db"""
    request = DwdObservationRequest(
        parameter=DwdObservationDataset.CLIMATE_SUMMARY,
        resolution=DwdObservationResolution.DAILY,
        start_date="2019-01-01",
        end_date="2020-01-01",
        settings=settings_si_false_wide_shape,
    ).filter_by_station_id(
        station_id=[1048],
    )
    filename = tmp_path.joinpath("observation.sqlite")
    df = request.values.all().df
    ExportMixin(df=df).to_target(f"sqlite:///{filename}?table=testdrive")
    connection = sqlite3.connect(filename)
    cursor = connection.cursor()
    cursor.execute("SELECT * FROM testdrive")
    results = cursor.fetchall()
    cursor.close()
    connection.close()
    first = list(results[0])
    first[2] = dt.datetime.fromisoformat(first[2])
    assert first == [
        "01048",
        "climate_summary",
        dt.datetime(2019, 1, 1),
        19.9,
        10.0,
        8.5,
        10.0,
        0.9,
        10.0,
        8.0,
        10.0,
        0.0,
        10.0,
        0.0,
        10.0,
        7.4,
        10.0,
        7.9,
        10.0,
        991.9,
        10.0,
        5.9,
        10.0,
        84.0,
        10.0,
        7.5,
        10.0,
        2.0,
        10.0,
        1.5,
        10.0,
    ]
    last = list(results[-1])
    last[2] = dt.datetime.fromisoformat(last[2])
    assert last == [
        "01048",
        "climate_summary",
        dt.datetime(2020, 1, 1),
        6.9,
        10.0,
        3.2,
        10.0,
        0.0,
        10.0,
        0.0,
        10.0,
        3.9,
        10.0,
        0.0,
        10.0,
        4.2,
        10.0,
        5.7,
        10.0,
        1005.1,
        10.0,
        2.4,
        10.0,
        79.0,
        10.0,
        5.6,
        10.0,
        -2.8,
        10.0,
        -4.6,
        10.0,
    ]


@pytest.mark.remote
def test_export_cratedb(
    settings_si_false,
):
    """Test export of DataFrame to cratedb"""
    request = DwdObservationRequest(
        parameter=DwdObservationDataset.CLIMATE_SUMMARY,
        resolution=DwdObservationResolution.DAILY,
        period=DwdObservationPeriod.RECENT,
        settings=settings_si_false,
    ).filter_by_station_id(
        station_id=[1048],
    )
    with mock.patch(
        "pandas.DataFrame.to_sql",
    ) as mock_to_sql:
        df = request.values.all().df
        ExportMixin(df=df).to_target("crate://localhost/?database=test&table=testdrive")
        mock_to_sql.assert_called_once_with(
            name="testdrive",
            con="crate://localhost",
            schema="test",
            if_exists="replace",
            index=False,
            chunksize=5000,
        )


@pytest.mark.remote
def test_export_duckdb(settings_si_false, tmp_path):
    """Test export of DataFrame to duckdb"""
    import duckdb

    request = DwdObservationRequest(
        parameter=DwdObservationDataset.CLIMATE_SUMMARY,
        resolution=DwdObservationResolution.DAILY,
        period=DwdObservationPeriod.HISTORICAL,
        settings=settings_si_false,
    ).filter_by_station_id(station_id=[1048])
    filename = tmp_path.joinpath("test.duckdb")
    df = request.values.all().df
    ExportMixin(df=df).to_target(f"duckdb:///{filename}?table=testdrive")
    connection = duckdb.connect(str(filename), read_only=True)
    cursor = connection.cursor()
    cursor.execute("SELECT * FROM testdrive")
    results = cursor.fetchall()
    cursor.close()
    connection.close()
    assert results[300_000] == (
        "01048",
        "climate_summary",
        "temperature_air_min_200",
        dt.datetime(1939, 7, 26),
        10.0,
        1.0,
    )


@surrogate("influxdb.InfluxDBClient")
@pytest.mark.remote
def test_export_influxdb1_tabular(settings_si_false_wide_shape):
    """Test export of DataFrame to influxdb v1"""
    request = DwdObservationRequest(
        parameter=DwdObservationDataset.CLIMATE_SUMMARY,
        resolution=DwdObservationResolution.DAILY,
        period=DwdObservationPeriod.RECENT,
        settings=settings_si_false_wide_shape,
    ).filter_by_station_id(station_id=[1048])
    mock_client = mock.MagicMock()
    with mock.patch(
        "influxdb.InfluxDBClient",
        side_effect=[mock_client],
        create=True,
    ) as mock_connect:
        df = request.values.all().df
        ExportMixin(df=df).to_target("influxdb://localhost/?database=dwd&table=weather")
        mock_connect.assert_called_once_with(
            host="localhost",
            port=8086,
            username=None,
            password=None,
            database="dwd",
            ssl=False,
        )
        mock_client.create_database.assert_called_once_with("dwd")
        mock_client.write_points.assert_called_once()
        mock_client.write_points.assert_called_with(
            points=mock.ANY,
            batch_size=50000,
        )
        points = mock_client.write_points.call_args.kwargs["points"]
        assert points[0]["measurement"] == "weather"
        assert list(points[0]["fields"].keys()) == [
            "station_id",
            "dataset",
            "wind_gust_max",
            "qn_wind_gust_max",
            "wind_speed",
            "qn_wind_speed",
            "precipitation_height",
            "qn_precipitation_height",
            "precipitation_form",
            "qn_precipitation_form",
            "sunshine_duration",
            "qn_sunshine_duration",
            "snow_depth",
            "qn_snow_depth",
            "cloud_cover_total",
            "qn_cloud_cover_total",
            "pressure_vapor",
            "qn_pressure_vapor",
            "pressure_air_site",
            "qn_pressure_air_site",
            "temperature_air_mean_200",
            "qn_temperature_air_mean_200",
            "humidity",
            "qn_humidity",
            "temperature_air_max_200",
            "qn_temperature_air_max_200",
            "temperature_air_min_200",
            "qn_temperature_air_min_200",
            "temperature_air_min_005",
            "qn_temperature_air_min_005",
        ]


@surrogate("influxdb.InfluxDBClient")
@pytest.mark.remote
def test_export_influxdb1_tidy(settings_si_false):
    """Test export of DataFrame to influxdb v1"""
    request = DwdObservationRequest(
        parameter=DwdObservationDataset.CLIMATE_SUMMARY,
        resolution=DwdObservationResolution.DAILY,
        period=DwdObservationPeriod.RECENT,
        settings=settings_si_false,
    ).filter_by_station_id(station_id=[1048])
    mock_client = mock.MagicMock()
    with mock.patch(
        "influxdb.InfluxDBClient",
        side_effect=[mock_client],
        create=True,
    ) as mock_connect:
        df = request.values.all().df
        ExportMixin(df=df).to_target("influxdb://localhost/?database=dwd&table=weather")
        mock_connect.assert_called_once_with(
            host="localhost",
            port=8086,
            username=None,
            password=None,
            database="dwd",
            ssl=False,
        )
        mock_client.create_database.assert_called_once_with("dwd")
        mock_client.write_points.assert_called_once()
        mock_client.write_points.assert_called_with(
            points=mock.ANY,
            batch_size=50000,
        )
        points = mock_client.write_points.call_args.kwargs["points"]
        assert points[0]["measurement"] == "weather"
        assert list(points[0]["fields"].keys()) == [
            "station_id",
            "dataset",
            "parameter",
            "value",
            "quality",
        ]


@surrogate("influxdb_client.InfluxDBClient")
@surrogate("influxdb_client.Point")
@surrogate("influxdb_client.client.write_api.SYNCHRONOUS")
@pytest.mark.remote
def test_export_influxdb2_tabular(settings_si_false):
    """Test export of DataFrame to influxdb v2"""
    request = DwdObservationRequest(
        parameter=DwdObservationDataset.CLIMATE_SUMMARY,
        resolution=DwdObservationResolution.DAILY,
        period=DwdObservationPeriod.RECENT,
        settings=settings_si_false,
    ).filter_by_station_id(station_id=[1048])
    mock_client = mock.MagicMock()
    with mock.patch(
        "influxdb_client.InfluxDBClient",
        side_effect=[mock_client],
        create=True,
    ) as mock_connect:
        with mock.patch(
            "influxdb_client.Point",
            create=True,
        ):
            df = request.values.all().df
            ExportMixin(df=df).to_target("influxdb2://orga:token@localhost/?database=dwd&table=weather")
            mock_connect.assert_called_once_with(url="http://localhost:8086", org="orga", token="token")  # noqa: S106


@surrogate("influxdb_client.InfluxDBClient")
@surrogate("influxdb_client.Point")
@surrogate("influxdb_client.client.write_api.SYNCHRONOUS")
@pytest.mark.remote
def test_export_influxdb2_tidy(settings_si_false):
    """Test export of DataFrame to influxdb v2"""
    request = DwdObservationRequest(
        parameter=DwdObservationDataset.CLIMATE_SUMMARY,
        resolution=DwdObservationResolution.DAILY,
        period=DwdObservationPeriod.RECENT,
        settings=settings_si_false,
    ).filter_by_station_id(station_id=[1048])
    mock_client = mock.MagicMock()
    with mock.patch(
        "influxdb_client.InfluxDBClient",
        side_effect=[mock_client],
        create=True,
    ) as mock_connect:
        with mock.patch(
            "influxdb_client.Point",
            create=True,
        ):
            df = request.values.all().df
            ExportMixin(df=df).to_target("influxdb2://orga:token@localhost/?database=dwd&table=weather")
            mock_connect.assert_called_once_with(url="http://localhost:8086", org="orga", token="token")  # noqa: S106


@surrogate("influxdb_client_3.InfluxDBClient3")
@surrogate("influxdb_client_3.Point")
@surrogate("influxdb_client_3.WriteOptions")
@surrogate("influxdb_client_3.write_client_options")
@surrogate("influxdb_client_3.write_client.client.write_api.WriteType")
@pytest.mark.remote
def test_export_influxdb3_tabular(settings_si_false):
    """Test export of DataFrame to influxdb v3"""
    request = DwdObservationRequest(
        parameter=DwdObservationDataset.CLIMATE_SUMMARY,
        resolution=DwdObservationResolution.DAILY,
        period=DwdObservationPeriod.RECENT,
        settings=settings_si_false,
    ).filter_by_station_id(station_id=[1048])
    with mock.patch(
        "influxdb_client_3.InfluxDBClient3",
    ) as mock_client, mock.patch(
        "influxdb_client_3.Point",
    ), mock.patch(
        "influxdb_client_3.WriteOptions",
    ) as mock_write_options, mock.patch(
        "influxdb_client_3.write_client_options",
    ) as mock_write_client_options, mock.patch(
        "influxdb_client_3.write_client.client.write_api.WriteType",
    ) as mock_write_type:
        df = request.values.all().df
        ExportMixin(df=df).to_target("influxdb3://orga:token@localhost/?database=dwd&table=weather")
        mock_write_options.assert_called_once_with(write_type=mock_write_type.synchronous)
        mock_write_client_options.assert_called_once_with(WriteOptions=mock_write_options())
        mock_client.assert_called_once_with(
            host="localhost",
            org="orga",
            database="dwd",
            token="token",  # noqa: S106
            write_client_options=mock_write_client_options(),
        )


@surrogate("influxdb_client_3.InfluxDBClient3")
@surrogate("influxdb_client_3.Point")
@surrogate("influxdb_client_3.WriteOptions")
@surrogate("influxdb_client_3.write_client_options")
@surrogate("influxdb_client_3.write_client.client.write_api.WriteType")
@pytest.mark.remote
def test_export_influxdb3_tidy(settings_si_false):
    """Test export of DataFrame to influxdb v3"""
    request = DwdObservationRequest(
        parameter=DwdObservationDataset.CLIMATE_SUMMARY,
        resolution=DwdObservationResolution.DAILY,
        period=DwdObservationPeriod.RECENT,
        settings=settings_si_false,
    ).filter_by_station_id(station_id=[1048])
    with mock.patch(
        "influxdb_client_3.InfluxDBClient3",
    ) as mock_client, mock.patch(
        "influxdb_client_3.Point",
    ), mock.patch(
        "influxdb_client_3.WriteOptions",
    ) as mock_write_options, mock.patch(
        "influxdb_client_3.write_client_options",
    ) as mock_write_client_options, mock.patch(
        "influxdb_client_3.write_client.client.write_api.WriteType",
    ) as mock_write_type:
        df = request.values.all().df
        ExportMixin(df=df).to_target("influxdb3://orga:token@localhost/?database=dwd&table=weather")
        mock_write_options.assert_called_once_with(write_type=mock_write_type.synchronous)
        mock_write_client_options.assert_called_once_with(WriteOptions=mock_write_options())
        mock_client.assert_called_once_with(
            host="localhost",
            org="orga",
            database="dwd",
            token="token",  # noqa: S106
            write_client_options=mock_write_client_options(),
        )
