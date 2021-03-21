# -*- coding: utf-8 -*-
# Copyright (c) 2018-2021, earthobservations developers.
# Distributed under the MIT License. See LICENSE for more info.
import datetime
import json
import os

import dateutil.parser
import mock
import pandas as pd
import pytest
from surrogate import surrogate

from wetterdienst.dwd.observations import (
    DwdObservationDataset,
    DwdObservationPeriod,
    DwdObservationRequest,
    DwdObservationResolution,
)
from wetterdienst.metadata.resolution import Resolution

df_station = pd.DataFrame.from_dict(
    {
        "STATION_ID": ["19087"],
        "FROM_DATE": [dateutil.parser.isoparse("1957-05-01T00:00:00.000Z")],
        "TO_DATE": [dateutil.parser.isoparse("1995-11-30T00:00:00.000Z")],
        "HEIGHT": [645.0],
        "LATITUDE": [48.8049],
        "LONGITUDE": [13.5528],
        "STATION_NAME": ["Freyung vorm Wald"],
        "STATE": ["Bayern"],
        "HAS_FILE": [False],
    }
)

df_data = pd.DataFrame.from_dict(
    {
        "STATION_ID": ["01048"],
        "DATASET": ["CLIMATE_SUMMARY"],
        "PARAMETER": ["TEMPERATURE_AIR_MAX_200"],
        "DATE": [dateutil.parser.isoparse("2019-12-28T00:00:00.000Z")],
        "VALUE": [1.3],
        "QUALITY": [None],
    }
)


def test_lowercase():

    df = df_data.dwd.lower()

    assert list(df.columns) == [
        "station_id",
        "dataset",
        "parameter",
        "date",
        "value",
        "quality",
    ]

    assert df.iloc[0]["dataset"] == "climate_summary"
    assert df.iloc[0]["parameter"] == "temperature_air_max_200"


def test_to_dict():
    df = df_data.io.to_dict()

    assert df == [
        {
            "DATASET": "CLIMATE_SUMMARY",
            "DATE": "2019-12-28T00:00:00+00:00",
            "PARAMETER": "TEMPERATURE_AIR_MAX_200",
            "QUALITY": None,
            "STATION_ID": "01048",
            "VALUE": 1.3,
        },
    ]


def test_filter_by_date():

    df = df_data.dwd.filter_by_date("2019-12-28", Resolution.HOURLY)
    assert not df.empty

    df = df_data.dwd.filter_by_date("2019-12-27", Resolution.HOURLY)
    assert df.empty


def test_filter_by_date_interval():

    df = df_data.dwd.filter_by_date("2019-12-27/2019-12-29", Resolution.HOURLY)
    assert not df.empty

    df = df_data.dwd.filter_by_date("2020/2022", Resolution.HOURLY)
    assert df.empty


def test_filter_by_date_monthly():

    result = pd.DataFrame.from_dict(
        {
            "STATION_ID": ["01048"],
            "DATASET": ["climate_summary"],
            "PARAMETER": ["temperature_air_max_200"],
            "FROM_DATE": [dateutil.parser.isoparse("2019-12-28T00:00:00.000Z")],
            "TO_DATE": [dateutil.parser.isoparse("2020-01-28T00:00:00.000Z")],
            "VALUE": [1.3],
            "QUALITY": [None],
        }
    )

    df = result.dwd.filter_by_date("2019-12/2020-01", Resolution.MONTHLY)
    assert not df.empty

    df = result.dwd.filter_by_date("2020/2022", Resolution.MONTHLY)
    assert df.empty

    df = result.dwd.filter_by_date("2020", Resolution.MONTHLY)
    assert df.empty


def test_filter_by_date_annual():

    result = pd.DataFrame.from_dict(
        {
            "STATION_ID": ["01048"],
            "DATASET": ["climate_summary"],
            "PARAMETER": ["temperature_air_max_200"],
            "FROM_DATE": [dateutil.parser.isoparse("2019-01-01T00:00:00.000Z")],
            "TO_DATE": [dateutil.parser.isoparse("2019-12-31T00:00:00.000Z")],
            "VALUE": [1.3],
            "QUALITY": [None],
        }
    )

    df = result.dwd.filter_by_date("2019-05/2019-09", Resolution.ANNUAL)
    assert not df.empty

    df = result.dwd.filter_by_date("2020/2022", Resolution.ANNUAL)
    assert df.empty

    df = result.dwd.filter_by_date("2020", Resolution.ANNUAL)
    assert df.empty


@pytest.mark.sql
def test_filter_by_sql():
    # TODO: change this to a test of historical data
    df = df_data.dwd.lower().io.sql(
        "SELECT * FROM data WHERE parameter='temperature_air_max_200' AND value < 1.5"
    )
    assert not df.empty

    df = df_data.dwd.lower().io.sql(
        "SELECT * FROM data WHERE parameter='temperature_air_max_200' AND value > 1.5"
    )
    assert df.empty


def test_format_json():

    output = df_data.dwd.lower().io.format("json")

    response = json.loads(output)
    station_ids = list(set([reading["station_id"] for reading in response]))

    assert "01048" in station_ids


def test_format_geojson():

    output = df_station.dwd.format("geojson")

    response = json.loads(output)

    station_names = [station["properties"]["name"] for station in response["features"]]

    assert "Freyung vorm Wald" in station_names


def test_format_csv():

    output = df_data.dwd.lower().io.format("csv").strip()

    assert "station_id,dataset,parameter,date,value,quality" in output
    assert (
        "01048,climate_summary,temperature_air_max_200,2019-12-28T00-00-00,1.3,"
        in output
    )


def test_format_unknown_dwd():

    with pytest.raises(KeyError) as ex:
        df_data.dwd.format("foobar")

    ex.match("Unknown output format")


def test_format_unknown_io():

    with pytest.raises(KeyError) as ex:
        df_data.io.format("foobar")

    ex.match("Unknown output format")


def test_request():

    request = DwdObservationRequest(
        parameter=DwdObservationDataset.CLIMATE_SUMMARY,
        resolution=DwdObservationResolution.DAILY,
        period=DwdObservationPeriod.RECENT,
    ).filter(station_id=[1048])

    df = request.values.all().df

    assert not df.empty


def test_export_unknown():

    request = DwdObservationRequest(
        parameter=DwdObservationDataset.CLIMATE_SUMMARY,
        resolution=DwdObservationResolution.DAILY,
        period=DwdObservationPeriod.RECENT,
    ).filter(
        station_id=[1048],
    )

    df = request.values.all().df

    with pytest.raises(KeyError) as ex:
        df.io.export("file:///test.foobar")

    ex.match("Unknown export file type")


def test_export_spreadsheet(tmpdir_factory):

    import openpyxl

    # 1. Request data and save to .xlsx file.
    request = DwdObservationRequest(
        parameter=DwdObservationDataset.CLIMATE_SUMMARY,
        resolution=DwdObservationResolution.DAILY,
        start_date="2019",
        end_date="2020",
        tidy_data=False,
    ).filter(
        station_id=[1048],
    )

    df = request.values.all().df

    filename = tmpdir_factory.mktemp("data").join("observations.xlsx")
    df.io.export(f"file://{filename}")

    workbook = openpyxl.load_workbook(filename=filename)
    worksheet = workbook.active

    # 2. Validate some details of .xlsx file.

    # Validate header row.
    header = list(worksheet.iter_cols(min_row=1, max_row=1, values_only=True))
    assert header == [
        ("DATE",),
        ("STATION_ID",),
        ("QN_3",),
        ("WIND_GUST_MAX",),
        ("WIND_SPEED",),
        ("QN_4",),
        ("PRECIPITATION_HEIGHT",),
        ("PRECIPITATION_FORM",),
        ("SUNSHINE_DURATION",),
        ("SNOW_DEPTH",),
        ("CLOUD_COVER_TOTAL",),
        ("PRESSURE_VAPOR",),
        ("PRESSURE_AIR",),
        ("TEMPERATURE_AIR_200",),
        ("HUMIDITY",),
        ("TEMPERATURE_AIR_MAX_200",),
        ("TEMPERATURE_AIR_MIN_200",),
        ("TEMPERATURE_AIR_MIN_005",),
    ]

    # Validate number of records.
    assert worksheet.max_row == 367

    first_record = list(worksheet.iter_cols(min_row=2, max_row=2, values_only=True))
    assert first_record == [
        ("2019-01-01T00:00:00+00:00",),
        ("01048",),
        (10,),
        (19.9,),
        (8.5,),
        (3,),
        (0.9,),
        (8,),
        (0,),
        (0,),
        (7.4,),
        (7.9,),
        (991.87,),
        (5.9,),
        (84.21,),
        (7.5,),
        (2,),
        (1.5,),
    ]

    last_record = list(
        worksheet.iter_cols(
            min_row=worksheet.max_row, max_row=worksheet.max_row, values_only=True
        )
    )
    assert last_record == [
        ("2020-01-01T00:00:00+00:00",),
        ("01048",),
        (10,),
        (6.9,),
        (3.2,),
        (3,),
        (0,),
        (0,),
        (3.933,),
        (0,),
        (4.2,),
        (5.7,),
        (1005.11,),
        (2.4,),
        (79,),
        (5.6,),
        (-2.8,),
        (-4.6,),
    ]

    os.unlink(filename)


def test_export_parquet(tmpdir_factory):

    import pyarrow.parquet as pq

    # Request data.
    request = DwdObservationRequest(
        parameter=DwdObservationDataset.CLIMATE_SUMMARY,
        resolution=DwdObservationResolution.DAILY,
        start_date="2019",
        end_date="2020",
        tidy_data=False,
    ).filter(
        station_id=[1048],
    )

    df = request.values.all().df

    # Save to Parquet file.
    filename = tmpdir_factory.mktemp("data").join("observations.parquet")
    df.io.export(f"file://{filename}")

    # Read back Parquet file.
    table = pq.read_table(filename)

    # Validate dimensions.
    assert table.num_columns == 18
    assert table.num_rows == 366

    # Validate column names.
    assert table.column_names == [
        "DATE",
        "STATION_ID",
        "QN_3",
        "WIND_GUST_MAX",
        "WIND_SPEED",
        "QN_4",
        "PRECIPITATION_HEIGHT",
        "PRECIPITATION_FORM",
        "SUNSHINE_DURATION",
        "SNOW_DEPTH",
        "CLOUD_COVER_TOTAL",
        "PRESSURE_VAPOR",
        "PRESSURE_AIR",
        "TEMPERATURE_AIR_200",
        "HUMIDITY",
        "TEMPERATURE_AIR_MAX_200",
        "TEMPERATURE_AIR_MIN_200",
        "TEMPERATURE_AIR_MIN_005",
    ]

    # Validate content.
    data = table.to_pydict()

    assert data["DATE"][0] == datetime.datetime(
        2019, 1, 1, 0, 0, tzinfo=datetime.timezone.utc
    )
    assert data["TEMPERATURE_AIR_MIN_005"][0] == 1.5
    assert data["DATE"][-1] == datetime.datetime(
        2020, 1, 1, 0, 0, tzinfo=datetime.timezone.utc
    )
    assert data["TEMPERATURE_AIR_MIN_005"][-1] == -4.6

    os.unlink(filename)


def test_export_feather(tmpdir_factory):

    import pyarrow.feather as feather

    # Request data
    request = DwdObservationRequest(
        parameter=DwdObservationDataset.CLIMATE_SUMMARY,
        resolution=DwdObservationResolution.DAILY,
        start_date="2019",
        end_date="2020",
        tidy_data=False,
    ).filter(
        station_id=[1048],
    )

    df = request.values.all().df

    # Save to Feather file.
    filename = tmpdir_factory.mktemp("data").join("observations.feather")
    df.io.export(f"file://{filename}")

    # Read back Feather file.
    table = feather.read_table(filename)

    # Validate dimensions.
    assert table.num_columns == 18
    assert table.num_rows == 366

    # Validate column names.
    assert table.column_names == [
        "DATE",
        "STATION_ID",
        "QN_3",
        "WIND_GUST_MAX",
        "WIND_SPEED",
        "QN_4",
        "PRECIPITATION_HEIGHT",
        "PRECIPITATION_FORM",
        "SUNSHINE_DURATION",
        "SNOW_DEPTH",
        "CLOUD_COVER_TOTAL",
        "PRESSURE_VAPOR",
        "PRESSURE_AIR",
        "TEMPERATURE_AIR_200",
        "HUMIDITY",
        "TEMPERATURE_AIR_MAX_200",
        "TEMPERATURE_AIR_MIN_200",
        "TEMPERATURE_AIR_MIN_005",
    ]

    # Validate content.
    data = table.to_pydict()

    assert data["DATE"][0] == datetime.datetime(
        2019, 1, 1, 0, 0, tzinfo=datetime.timezone.utc
    )
    assert data["TEMPERATURE_AIR_MIN_005"][0] == 1.5
    assert data["DATE"][-1] == datetime.datetime(
        2020, 1, 1, 0, 0, tzinfo=datetime.timezone.utc
    )
    assert data["TEMPERATURE_AIR_MIN_005"][-1] == -4.6

    os.unlink(filename)


def test_export_sqlite(tmpdir_factory):

    import sqlite3

    request = DwdObservationRequest(
        parameter=DwdObservationDataset.CLIMATE_SUMMARY,
        resolution=DwdObservationResolution.DAILY,
        start_date="2019",
        end_date="2020",
        tidy_data=False,
    ).filter(
        station_id=[1048],
    )

    filename = tmpdir_factory.mktemp("data").join("observations.sqlite")

    df = request.values.all().df
    df.io.export(f"sqlite:///{filename}?table=testdrive")

    connection = sqlite3.connect(filename)
    cursor = connection.cursor()
    cursor.execute("SELECT * FROM testdrive")
    results = cursor.fetchall()
    cursor.close()
    connection.close()

    assert results[0] == (
        "2019-01-01 00:00:00.000000",
        "01048",
        10,
        19.9,
        8.5,
        3,
        0.9,
        8,
        0.0,
        0,
        7.4,
        7.9,
        991.87,
        5.9,
        84.21,
        7.5,
        2.0,
        1.5,
    )

    assert results[-1] == (
        "2020-01-01 00:00:00.000000",
        "01048",
        10,
        6.9,
        3.2,
        3,
        0.0,
        0,
        3.933,
        0,
        4.2,
        5.7,
        1005.11,
        2.4,
        79.0,
        5.6,
        -2.8,
        -4.6,
    )


def test_export_cratedb():

    request = DwdObservationRequest(
        parameter=DwdObservationDataset.CLIMATE_SUMMARY,
        resolution=DwdObservationResolution.DAILY,
        period=DwdObservationPeriod.RECENT,
    ).filter(
        station_id=[1048],
    )

    with mock.patch(
        "pandas.DataFrame.to_sql",
    ) as mock_to_sql:

        df = request.values.all().df
        df.dwd.lower().io.export("crate://localhost/?database=test&table=testdrive")

        mock_to_sql.assert_called_once_with(
            name="testdrive",
            con="crate://localhost",
            schema="test",
            if_exists="replace",
            index=False,
            # method="multi",
            chunksize=5000,
        )


@surrogate("duckdb.connect")
def test_export_duckdb():

    request = DwdObservationRequest(
        parameter=DwdObservationDataset.CLIMATE_SUMMARY,
        resolution=DwdObservationResolution.DAILY,
        period=DwdObservationPeriod.RECENT,
    ).filter(station_id=[1048])

    mock_connection = mock.MagicMock()
    with mock.patch(
        "duckdb.connect", side_effect=[mock_connection], create=True
    ) as mock_connect:

        df = request.values.all().df
        df.io.export("duckdb:///test.duckdb?table=testdrive")

        mock_connect.assert_called_once_with(database="test.duckdb", read_only=False)
        mock_connection.register.assert_called_once()
        mock_connection.execute.assert_called()
        mock_connection.table.assert_called_once_with("testdrive")
        # a.table.to_df.assert_called()
        mock_connection.close.assert_called_once()


@surrogate("influxdb.dataframe_client.DataFrameClient")
def test_export_influxdb_tabular():

    request = DwdObservationRequest(
        parameter=DwdObservationDataset.CLIMATE_SUMMARY,
        resolution=DwdObservationResolution.DAILY,
        period=DwdObservationPeriod.RECENT,
        tidy_data=False,
    ).filter(station_id=[1048])

    mock_client = mock.MagicMock()
    with mock.patch(
        "influxdb.dataframe_client.DataFrameClient",
        side_effect=[mock_client],
        create=True,
    ) as mock_connect:

        df = request.values.all().df
        df.dwd.lower().io.export("influxdb://localhost/?database=dwd&table=weather")

        mock_connect.assert_called_once_with(database="dwd")
        mock_client.create_database.assert_called_once_with("dwd")
        mock_client.write_points.assert_called_once()

        mock_client.write_points.assert_called_with(
            dataframe=mock.ANY,
            measurement="weather",
            tag_columns=["station_id", "qn_3", "qn_4"],
            batch_size=50000,
        )


@surrogate("influxdb.dataframe_client.DataFrameClient")
def test_export_influxdb_tidy():

    request = DwdObservationRequest(
        parameter=DwdObservationDataset.CLIMATE_SUMMARY,
        resolution=DwdObservationResolution.DAILY,
        period=DwdObservationPeriod.RECENT,
        tidy_data=True,
    ).filter(station_id=[1048])

    mock_client = mock.MagicMock()
    with mock.patch(
        "influxdb.dataframe_client.DataFrameClient",
        side_effect=[mock_client],
        create=True,
    ) as mock_connect:

        df = request.values.all().df
        df.dwd.lower().io.export("influxdb://localhost/?database=dwd&table=weather")

        mock_connect.assert_called_once_with(database="dwd")
        mock_client.create_database.assert_called_once_with("dwd")
        mock_client.write_points.assert_called_once()

        mock_client.write_points.assert_called_with(
            dataframe=mock.ANY,
            measurement="weather",
            tag_columns=["station_id", "quality", "dataset", "parameter"],
            batch_size=50000,
        )
